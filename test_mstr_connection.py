import requests
import json
import urllib3
import time
import os
import sys
import glob
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment

# Deshabilitar advertencias de certificados SSL no verificados
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configuración base de MicroStrategy
BASE_URL = "https://env-i921eu432wwh4k73.cloud.strategy.com/MicroStrategyLibrary"
API_URL = f"{BASE_URL}/api"
DOSSIER_ID = "06C17B674C66C15648D532B59505E1E3"  # ID de la conexión de Reporte

# ==============================================================================
# POR FAVOR INGRESA TUS CREDENCIALES DE MICROSTRATEGY AQUÍ:
# ==============================================================================
USERNAME = "maria.sanchez"
PASSWORD = "Marzo0393*"
# ==============================================================================

# Resolver directorios locales
try:
    dir_path = os.path.dirname(os.path.abspath(__file__))
except NameError:
    dir_path = r"c:\Users\Maria Alejandra\OneDrive - VIRTUALSOFT SERVICIOS & SOFTWARE S.A.S\Antigravity\Retiros"

excel_out = os.path.join(dir_path, "Reporte_Efectividad_Automatizacion.xlsx")
html_out = os.path.join(dir_path, "Dashboard_Efectividad.html")

def clean_nan(obj):
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan(v) for v in obj]
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    return obj

def parse_mstr_node(node, attribute_names, current_row_data=None):
    if current_row_data is None:
        current_row_data = {}
    current_row_data = current_row_data.copy()
    
    if "element" in node:
        element = node["element"]
        attr_idx = element.get("attributeIndex")
        if attr_idx is not None and attr_idx < len(attribute_names):
            attr_name = attribute_names[attr_idx]
            current_row_data[attr_name] = element.get("name")
            
    if "metrics" in node:
        metrics = node["metrics"]
        for metric_name, metric_val in metrics.items():
            current_row_data[metric_name] = metric_val.get("rv")
            
    rows = []
    if "children" in node:
        for child in node["children"]:
            rows.extend(parse_mstr_node(child, attribute_names, current_row_data))
    else:
        rows.append(current_row_data)
        
    return rows

def clean_mstr_dataframe(df):
    print("    [*] Aplicando reglas de limpieza sobre la data...")
    initial_rows = len(df)
    
    # Regla 1: Construir/Normalizar columna Marca
    if 'Marca' not in df.columns or df['Marca'].isna().all():
        df['Marca'] = df['Partner'].astype(str).str.strip() + ' ' + df['Pais'].astype(str).str.strip()
        cols = ['Marca'] + [col for col in df.columns if col not in ['Partner', 'Pais', 'Marca']]
        df_clean = df[cols].copy()
    else:
        df_clean = df.copy()

    # Regla 2: Excluir estados no procesados (Pendiente por Pago y Activo sin completar)
    # y excluidos según el modelo de análisis manual
    mask_excl = (df_clean['Estado Retiro Creado'].isin(['Pendiente por Pago', 'Activo'])) & (df_clean['Medio De Pago'] == 'Fisicamente')
    rows_before = len(df_clean)
    df_clean = df_clean[~mask_excl].copy()
    print(f"        - Filas excluidas por estado pendiente/físicamente: {rows_before - len(df_clean)}")
    
    # Regla 3: Eliminar duplicados por Id Retiro (conservando la última versión)
    rows_before_dup = len(df_clean)
    df_clean = df_clean.drop_duplicates(subset=['Id Retiro'], keep='last')
    print(f"        - Filas duplicadas por Id Retiro eliminadas: {rows_before_dup - len(df_clean)}")
    
    print(f"    [+] Limpieza completada. Filas finales limpias: {len(df_clean)}")
    return df_clean

# Dict of business rules limits by Brand and Risk level for simulation
RULES = {
    'Camanbet VES': {
        'Bajo': {'min': 1000, 'max': 20000},
        'Medio': {'min': 1000, 'max': 10000}
    },
    'Camanbet Venezuela': {
        'Bajo': {'min': 1000, 'max': 20000},
        'Medio': {'min': 1000, 'max': 10000}
    },
    'Camanbet Venezuela 2': {
        'Bajo': {'min': 1000, 'max': 20000},
        'Medio': {'min': 1000, 'max': 10000}
    },
    'Doradobet Chile': {
        'Bajo': {'min': 10000, 'max': 999999},
        'Medio': {'min': 10000, 'max': 500000}
    },
    'Doradobet Costa Rica': {
        'Bajo': {'min': 0, 'max': 249999},
        'Medio': {'min': 0, 'max': 249000}
    },
    'Doradobet Ecuador': {
        'Bajo': {'min': 0, 'max': 249},
        'Medio': {'min': 30, 'max': 250},
        'Sin Nivel': {'min': 0, 'max': 200}
    },
    'Ecuabet Ecuador': {
        'Bajo': {'min': 0, 'max': 249},
        'Medio': {'min': 30, 'max': 250},
        'Sin Nivel': {'min': 0, 'max': 200}
    },
    'Doradobet Guatemala': {
        'Sin Nivel': {'min': 100, 'max': 2000},
        'Bajo': {'min': 10, 'max': 5000},
        'Medio': {'min': 10, 'max': 5000}
    },
    'Doradobet Honduras': {
        'Sin Nivel': {'min': 0, 'max': 5000},
        'Bajo': {'min': 10, 'max': 10000},
        'Medio': {'min': 10, 'max': 10000}
    },
    'Doradobet Perú': {
        'Sin Nivel': {'min': 0, 'max': 999},
        'Bajo': {'min': 50, 'max': 5000},
        'Medio': {'min': 50, 'max': 1999}
    },
    'Doradobet Salvador': {
        'Sin Nivel': {'min': 10, 'max': 200},
        'Bajo': {'min': 10, 'max': 499},
        'Medio': {'min': 10, 'max': 500}
    },
    'Doradobet El Salvador': {
        'Sin Nivel': {'min': 10, 'max': 200},
        'Bajo': {'min': 10, 'max': 499},
        'Medio': {'min': 10, 'max': 500}
    },
    'Paniplay': {
        'Sin Nivel': {'min': 0, 'max': 5000},
        'Bajo': {'min': 10, 'max': 10000},
        'Medio': {'min': 10, 'max': 10000}
    },
    'Ganaplay Salvador': {
        'Bajo': {'min': 10, 'max': 300},
        'Medio': {'min': 10, 'max': 150}
    },
    'Ganaplay Guatemala': {
        'Bajo': {'min': 50, 'max': 2000},
        'Medio': {'min': 50, 'max': 1500}
    }
}

def prepare_dataframe(df):
    df = df.copy()
    
    # Convirtiendo a datetime (Optimizado con format='mixed' y errors='coerce')
    df['t_creacion'] = pd.to_datetime(df['Fecha Creacion Retiro Time'], format='mixed', errors='coerce')
    df['t_cambio'] = pd.to_datetime(df['Fecha Cambio Time'], format='mixed', errors='coerce')
    df['t_pago'] = pd.to_datetime(df['Fecha Pago Time'], format='mixed', errors='coerce')
    
    # Calcular diferencias en minutos
    df['Minutos Creado a Aprobado'] = (df['t_cambio'] - df['t_creacion']).dt.total_seconds() / 60.0
    df['Minutos Creado a Pagado'] = (df['t_pago'] - df['t_creacion']).dt.total_seconds() / 60.0
    df['Minutos Aprobado a Pagado'] = (df['t_pago'] - df['t_cambio']).dt.total_seconds() / 60.0
    
    # Reemplazar valores negativos por 0 por seguridad
    df.loc[df['Minutos Creado a Aprobado'] < 0, 'Minutos Creado a Aprobado'] = 0
    df.loc[df['Minutos Creado a Pagado'] < 0, 'Minutos Creado a Pagado'] = 0
    df.loc[df['Minutos Aprobado a Pagado'] < 0, 'Minutos Aprobado a Pagado'] = 0
    
    # Horas
    df['Horas Creado a Aprobado'] = df['Minutos Creado a Aprobado'] / 60.0
    df['Horas Creado a Pagado'] = df['Minutos Creado a Pagado'] / 60.0
    df['Horas Aprobado a Pagado'] = df['Minutos Aprobado a Pagado'] / 60.0
    
    # Normalizar Tipo Aprobación
    is_auto = df['Tipo Aprobacion'].astype(str).str.strip().str.lower().isin(['automatico', 'automático'])
    df['Tipo Aprobacion_Norm'] = np.where(is_auto, 'Automatico', 'Manual')
    
    # Normalizar Nivel De Riesgo
    df['Nivel De Riesgo_Norm'] = df['Nivel De Riesgo'].fillna('Sin Nivel')
    df['Nivel De Riesgo_Norm'] = df['Nivel De Riesgo_Norm'].replace({'nan': 'Sin Nivel', 'None': 'Sin Nivel', '': 'Sin Nivel'})
    df['Nivel De Riesgo_Norm'] = df['Nivel De Riesgo_Norm'].astype(str).str.strip()
    
    # Normalizar Marca
    df['Marca'] = df['Marca'].astype(str).str.strip()
    
    # Calculate simulated auto based on user spreadsheet (Automate all Bajo Risk manual)
    def is_sim_auto_row(row):
        if row['Tipo Aprobacion_Norm'] == 'Automatico':
            return True
        if row['Nivel De Riesgo_Norm'] == 'Bajo':
            return True
        return False
        
    df['Is_Simulated_Auto'] = df.apply(is_sim_auto_row, axis=1)
    
    return df

def analyze_month(df):
    total_retiros = int(len(df))
    is_auto = df['Tipo Aprobacion_Norm'] == 'Automatico'
    is_manual = df['Tipo Aprobacion_Norm'] == 'Manual'
    retiros_automaticos = int(is_auto.sum())
    pct_automatizacion = float((retiros_automaticos / total_retiros) * 100) if total_retiros > 0 else 0.0
    
    df_auto = df[is_auto]
    df_manual = df[is_manual]
    
    global_stats = {
        "total_retiros": total_retiros,
        "retiros_automaticos": retiros_automaticos,
        "pct_automatizacion": pct_automatizacion,
        "auto_times": {
            "creado_aprobado_mean_min": float(df_auto['Minutos Creado a Aprobado'].mean()) if len(df_auto) > 0 else 0.0,
            "creado_aprobado_median_min": float(df_auto['Minutos Creado a Aprobado'].median()) if len(df_auto) > 0 else 0.0,
            "creado_aprobado_mean_hrs": float(df_auto['Horas Creado a Aprobado'].mean()) if len(df_auto) > 0 else 0.0,
            "creado_aprobado_median_hrs": float(df_auto['Horas Creado a Aprobado'].median()) if len(df_auto) > 0 else 0.0,
            "creado_pago_mean_min": float(df_auto['Minutos Creado a Pagado'].mean()) if len(df_auto) > 0 else 0.0,
            "creado_pago_median_min": float(df_auto['Minutos Creado a Pagado'].median()) if len(df_auto) > 0 else 0.0,
            "creado_pago_mean_hrs": float(df_auto['Horas Creado a Pagado'].mean()) if len(df_auto) > 0 else 0.0,
            "creado_pago_median_hrs": float(df_auto['Horas Creado a Pagado'].median()) if len(df_auto) > 0 else 0.0,
            "aprobado_pago_mean_min": float(df_auto['Minutos Aprobado a Pagado'].mean()) if len(df_auto) > 0 else 0.0,
            "aprobado_pago_median_min": float(df_auto['Minutos Aprobado a Pagado'].median()) if len(df_auto) > 0 else 0.0,
            "aprobado_pago_mean_hrs": float(df_auto['Horas Aprobado a Pagado'].mean()) if len(df_auto) > 0 else 0.0,
            "aprobado_pago_median_hrs": float(df_auto['Horas Aprobado a Pagado'].median()) if len(df_auto) > 0 else 0.0
        },
        "manual_times": {
            "creado_aprobado_mean_min": float(df_manual['Minutos Creado a Aprobado'].mean()) if len(df_manual) > 0 else 0.0,
            "creado_aprobado_median_min": float(df_manual['Minutos Creado a Aprobado'].median()) if len(df_manual) > 0 else 0.0,
            "creado_aprobado_mean_hrs": float(df_manual['Horas Creado a Aprobado'].mean()) if len(df_manual) > 0 else 0.0,
            "creado_aprobado_median_hrs": float(df_manual['Horas Creado a Aprobado'].median()) if len(df_manual) > 0 else 0.0,
            "creado_pago_mean_min": float(df_manual['Minutos Creado a Pagado'].mean()) if len(df_manual) > 0 else 0.0,
            "creado_pago_median_min": float(df_manual['Minutos Creado a Pagado'].median()) if len(df_manual) > 0 else 0.0,
            "creado_pago_mean_hrs": float(df_manual['Horas Creado a Pagado'].mean()) if len(df_manual) > 0 else 0.0,
            "creado_pago_median_hrs": float(df_manual['Horas Creado a Pagado'].median()) if len(df_manual) > 0 else 0.0,
            "aprobado_pago_mean_min": float(df_manual['Minutos Aprobado a Pagado'].mean()) if len(df_manual) > 0 else 0.0,
            "aprobado_pago_median_min": float(df_manual['Minutos Aprobado a Pagado'].median()) if len(df_manual) > 0 else 0.0,
            "aprobado_pago_mean_hrs": float(df_manual['Horas Aprobado a Pagado'].mean()) if len(df_manual) > 0 else 0.0,
            "aprobado_pago_median_hrs": float(df_manual['Horas Aprobado a Pagado'].median()) if len(df_manual) > 0 else 0.0
        }
    }
    
    # Resumen de Marcas
    brands = sorted(df['Marca'].dropna().unique())
    brands_summary_list = []
    brand_rows_excel = []
    
    for brand in brands:
        df_b = df[df['Marca'] == brand]
        b_total = len(df_b)
        b_auto = df_b['Tipo Aprobacion_Norm'].eq('Automatico').sum()
        b_pct = (b_auto / b_total * 100) if b_total > 0 else 0.0
        
        # Simulated metrics
        b_sim = df_b['Is_Simulated_Auto'].sum()
        b_sim_pct = (b_sim / b_total * 100) if b_total > 0 else 0.0
        
        brands_summary_list.append({
            "Marca": brand,
            "Total Retiros": int(b_total),
            "Retiros Automaticos": int(b_auto),
            "% Automatizacion": float(b_pct),
            "Retiros Automaticos Proyectados": int(b_sim),
            "% Automatizacion Proyectado": float(b_sim_pct)
        })
        brand_rows_excel.append([
            brand, int(b_total), int(b_auto), float(b_pct),
            int(b_sim), float(b_sim_pct)
        ])
        
    df_brands_excel = pd.DataFrame(brand_rows_excel, columns=[
        'Marca', 'Total Retiros', 'Retiros Automaticos', '% Automatizacion',
        'Retiros Automaticos Proyectados', '% Automatizacion Proyectado'
    ])
    
    # Tabla de Detalles
    details_list = []
    detail_rows_excel = []
    for brand in brands:
        df_b = df[df['Marca'] == brand]
        for ta in ['Automatico', 'Manual']:
            df_bt = df_b[df_b['Tipo Aprobacion_Norm'] == ta]
            bt_total = len(df_bt)
            
            mean_ca = df_bt['Minutos Creado a Aprobado'].mean()
            med_ca = df_bt['Minutos Creado a Aprobado'].median()
            mean_ca_h = df_bt['Horas Creado a Aprobado'].mean()
            med_ca_h = df_bt['Horas Creado a Aprobado'].median()
            
            mean_cp = df_bt['Minutos Creado a Pagado'].mean()
            med_cp = df_bt['Minutos Creado a Pagado'].median()
            mean_cp_h = df_bt['Horas Creado a Pagado'].mean()
            med_cp_h = df_bt['Horas Creado a Pagado'].median()
            
            mean_ap = df_bt['Minutos Aprobado a Pagado'].mean()
            med_ap = df_bt['Minutos Aprobado a Pagado'].median()
            mean_ap_h = df_bt['Horas Aprobado a Pagado'].mean()
            med_ap_h = df_bt['Horas Aprobado a Pagado'].median()
            
            detail_dict = {
                "Marca": brand,
                "Tipo Aprobacion": ta,
                "Total Retiros": int(bt_total),
                "Promedio Creado a Aprobado (Min)": float(mean_ca) if pd.notna(mean_ca) else None,
                "Mediana Creado a Aprobado (Min)": float(med_ca) if pd.notna(med_ca) else None,
                "Promedio Creado a Aprobado (Hrs)": float(mean_ca_h) if pd.notna(mean_ca_h) else None,
                "Mediana Creado a Aprobado (Hrs)": float(med_ca_h) if pd.notna(med_ca_h) else None,
                "Promedio Creado a Pagado (Min)": float(mean_cp) if pd.notna(mean_cp) else None,
                "Mediana Creado a Pagado (Min)": float(med_cp) if pd.notna(med_cp) else None,
                "Promedio Creado a Pagado (Hrs)": float(mean_cp_h) if pd.notna(mean_cp_h) else None,
                "Mediana Creado a Pagado (Hrs)": float(med_cp_h) if pd.notna(med_cp_h) else None,
                "Promedio Aprobado a Pagado (Min)": float(mean_ap) if pd.notna(mean_ap) else None,
                "Mediana Aprobado a Pagado (Min)": float(med_ap) if pd.notna(med_ap) else None,
                "Promedio Aprobado a Pagado (Hrs)": float(mean_ap_h) if pd.notna(mean_ap_h) else None,
                "Mediana Aprobado a Pagado (Hrs)": float(med_ap_h) if pd.notna(med_ap_h) else None,
                "Tipo Aprobación": ta
            }
            details_list.append(detail_dict)
            
            detail_rows_excel.append([
                brand, ta, int(bt_total),
                mean_ca if pd.notna(mean_ca) else None,
                med_ca if pd.notna(med_ca) else None,
                mean_ca_h if pd.notna(mean_ca_h) else None,
                med_ca_h if pd.notna(med_ca_h) else None,
                mean_cp if pd.notna(mean_cp) else None,
                med_cp if pd.notna(med_cp) else None,
                mean_cp_h if pd.notna(mean_cp_h) else None,
                med_cp_h if pd.notna(med_cp_h) else None,
                mean_ap if pd.notna(mean_ap) else None,
                med_ap if pd.notna(med_ap) else None,
                mean_ap_h if pd.notna(mean_ap_h) else None,
                med_ap_h if pd.notna(med_ap_h) else None
            ])
            
    df_details_excel = pd.DataFrame(detail_rows_excel, columns=[
        'Marca', 'Tipo Aprobacion', 'Total Retiros',
        'Promedio Creado a Aprobado (Min)', 'Mediana Creado a Aprobado (Min)',
        'Promedio Creado a Aprobado (Hrs)', 'Mediana Creado a Aprobado (Hrs)',
        'Promedio Creado a Pagado (Min)', 'Mediana Creado a Pagado (Min)',
        'Promedio Creado a Pagado (Hrs)', 'Mediana Creado a Pagado (Hrs)',
        'Promedio Aprobado a Pagado (Min)', 'Mediana Aprobado a Pagado (Min)',
        'Promedio Aprobado a Pagado (Hrs)', 'Mediana Aprobado a Pagado (Hrs)'
    ])
    
    # Tabla Global de Riesgo
    risk_levels = ['Alto', 'Bajo', 'Medio', 'Sin Nivel']
    global_risk_list = []
    global_risk_rows_excel = []
    for rl in risk_levels:
        df_r = df[df['Nivel De Riesgo_Norm'] == rl]
        r_total = len(df_r)
        r_auto = df_r['Tipo Aprobacion_Norm'].eq('Automatico').sum()
        r_pct = (r_auto / r_total * 100) if r_total > 0 else 0.0
        
        val_auto = df_r[df_r['Tipo Aprobacion_Norm'] == 'Automatico']['Valor Retiros Creados'].mean()
        val_manual = df_r[df_r['Tipo Aprobacion_Norm'] == 'Manual']['Valor Retiros Creados'].mean()
        
        global_risk_list.append({
            "Nivel De Riesgo": rl,
            "Total Retiros": int(r_total),
            "Retiros Automaticos": int(r_auto),
            "% Automatizacion": float(r_pct),
            "Valor Promedio Automatico": float(val_auto) if pd.notna(val_auto) else 0.0,
            "Valor Promedio Manual": float(val_manual) if pd.notna(val_manual) else 0.0
        })
        global_risk_rows_excel.append([
            rl, int(r_total), int(r_auto), float(r_pct),
            float(val_auto) if pd.notna(val_auto) else None,
            float(val_manual) if pd.notna(val_manual) else None
        ])
        
    df_global_risk_excel = pd.DataFrame(global_risk_rows_excel, columns=[
        'Nivel De Riesgo', 'Total Retiros', 'Retiros Automaticos', '% Automatizacion',
        'Valor Promedio Automatico', 'Valor Promedio Manual'
    ])
    
    # Tabla de Riesgo por Marca
    brand_risk_list = []
    for brand in brands:
        df_b = df[df['Marca'] == brand]
        for rl in risk_levels:
            df_br = df_b[df_b['Nivel De Riesgo_Norm'] == rl]
            br_total = len(df_br)
            br_auto = df_br['Tipo Aprobacion_Norm'].eq('Automatico').sum()
            br_pct = (br_auto / br_total * 100) if br_total > 0 else 0.0
            
            val_auto = df_br[df_br['Tipo Aprobacion_Norm'] == 'Automatico']['Valor Retiros Creados'].mean()
            val_manual = df_br[df_br['Tipo Aprobacion_Norm'] == 'Manual']['Valor Retiros Creados'].mean()
            
            brand_risk_list.append({
                "Marca": brand,
                "Nivel De Riesgo": rl,
                "Total Retiros": int(br_total),
                "Retiros Automaticos": int(br_auto),
                "% Automatizacion": float(br_pct),
                "Valor Promedio Automatico": float(val_auto) if pd.notna(val_auto) else 0.0,
                "Valor Promedio Manual": float(val_manual) if pd.notna(val_manual) else 0.0
            })
            
    # Generar agregaciones analíticas
    df['date_str'] = df['t_cambio'].dt.strftime('%Y-%m-%d')
    agg_daily = df.groupby(['date_str', 'Marca', 'Nivel De Riesgo_Norm', 'Tipo Aprobacion_Norm']).agg(
        count=('Id Retiro', 'count'),
        sum_valor=('Valor Retiros Creados', 'sum'),
        sum_ca_min=('Minutos Creado a Aprobado', 'sum'),
        count_ca=('Minutos Creado a Aprobado', 'count'),
        sum_cp_min=('Minutos Creado a Pagado', 'sum'),
        count_cp=('Minutos Creado a Pagado', 'count'),
        sum_ap_min=('Minutos Aprobado a Pagado', 'sum'),
        count_ap=('Minutos Aprobado a Pagado', 'count'),
        auto_sim_count=('Is_Simulated_Auto', 'sum')
    ).reset_index()
    
    df['hour'] = df['t_creacion'].dt.hour
    agg_hourly = df.groupby(['hour', 'Marca', 'Nivel De Riesgo_Norm', 'Tipo Aprobacion_Norm']).agg(
        count=('Id Retiro', 'count'),
        sum_ca_min=('Minutos Creado a Aprobado', 'sum'),
        count_ca=('Minutos Creado a Aprobado', 'count'),
        auto_sim_count=('Is_Simulated_Auto', 'sum')
    ).reset_index()
    
    df['Banco_Norm'] = df['Nombre Banco'].fillna('Sin Banco').astype(str).str.strip()
    df.loc[df['Banco_Norm'] == '', 'Banco_Norm'] = 'Sin Banco'
    df.loc[df['Banco_Norm'].str.lower() == 'nan', 'Banco_Norm'] = 'Sin Banco'
    df.loc[df['Banco_Norm'].str.lower() == 'none', 'Banco_Norm'] = 'Sin Banco'
    agg_bank = df.groupby(['Banco_Norm', 'Marca', 'Nivel De Riesgo_Norm', 'Tipo Aprobacion_Norm']).agg(
        count=('Id Retiro', 'count'),
        sum_ap_min=('Minutos Aprobado a Pagado', 'sum'),
        count_ap=('Minutos Aprobado a Pagado', 'count'),
        auto_sim_count=('Is_Simulated_Auto', 'sum')
    ).reset_index()
    
    daily_list = agg_daily.to_dict(orient='records')
    hourly_list = agg_hourly.to_dict(orient='records')
    bank_list = agg_bank.to_dict(orient='records')
    
    json_data = {
        "global": global_stats,
        "brands_summary": brands_summary_list,
        "details": details_list,
        "global_risk": global_risk_list,
        "brand_risk": brand_risk_list,
        "daily_data": daily_list,
        "hourly_data": hourly_list,
        "bank_data": bank_list
    }
    
    dfs = {
        "brands_summary": df_brands_excel,
        "details": df_details_excel,
        "global_risk": df_global_risk_excel
    }
    
    return json_data, dfs

def find_month_files():
    pattern = os.path.join(dir_path, "Retiros *.xlsx")
    files = glob.glob(pattern)
    
    month_map = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
    }
    
    detected = []
    for f in files:
        name = os.path.basename(f)
        if any(x in name.lower() for x in ["backup", "prev", "raw"]):
            continue
        
        match = re.match(r"Retiros\s+([A-Za-z]+)\.xlsx", name, re.IGNORECASE)
        if match:
            m_name = match.group(1).lower()
            m_num = month_map.get(m_name, 99)
            detected.append({
                "month_name": m_name,
                "month_num": m_num,
                "file_path": f,
                "display_name": m_name.capitalize()
            })
    
    detected.sort(key=lambda x: x["month_num"])
    return detected

def run_reporting_pipeline():
    print("\n[*] Iniciando la tubería de generación de reportes...")
    month_files = find_month_files()
    if not month_files:
        print("[-] ERROR: No se encontraron archivos de meses (Retiros <Mes>.xlsx) en la carpeta.")
        return False
        
    print(f"[+] Archivos de meses detectados y ordenados: {[m['display_name'] for m in month_files]}")
    
    processed_months = []
    seen_ids = set()
    for m in month_files:
        excel_path = m['file_path']
        csv_path = excel_path.replace(".xlsx", "_cache.csv")
        
        # Intentar cargar desde el archivo CSV cacheado si existe y es más reciente que el archivo de Excel
        use_cache = False
        if os.path.exists(csv_path):
            excel_mtime = os.path.getmtime(excel_path)
            csv_mtime = os.path.getmtime(csv_path)
            if csv_mtime > excel_mtime:
                use_cache = True
                
        if use_cache:
            print(f"\n=== Procesando data de {m['display_name']} desde caché rápido '{os.path.basename(csv_path)}' ===")
            df_raw = pd.read_csv(csv_path)
        else:
            print(f"\n=== Procesando data de {m['display_name']} desde Excel '{os.path.basename(excel_path)}' ===")
            xl = pd.ExcelFile(excel_path)
            sheet = 'Retiros BD Conexion' if 'Retiros BD Conexion' in xl.sheet_names else xl.sheet_names[0]
            df_raw = xl.parse(sheet)
            try:
                df_raw.to_csv(csv_path, index=False)
                print(f"        [+] Caché rápido '{os.path.basename(csv_path)}' creado con éxito.")
            except Exception as e:
                print(f"        [-] Advertencia al crear caché CSV: {e}")
        
        # Deduplicación cronológica entre meses
        if 'Id Retiro' in df_raw.columns:
            df_raw['Id Retiro_str'] = df_raw['Id Retiro'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
            if len(seen_ids) > 0:
                rows_before = len(df_raw)
                df_raw = df_raw[~df_raw['Id Retiro_str'].isin(seen_ids)].copy()
                print(f"        - Excluidos {rows_before - len(df_raw)} registros repetidos en meses históricos anteriores.")
            seen_ids.update(df_raw['Id Retiro_str'].dropna())
            df_raw = df_raw.drop(columns=['Id Retiro_str'])
            
        df = prepare_dataframe(df_raw)
        m_json, m_dfs = analyze_month(df)
        processed_months.append({
            "month_name": m["month_name"],
            "display_name": m["display_name"],
            "json_data": m_json,
            "dfs": m_dfs
        })
        
    print("\n=== Calculando Resumen Comparativo de Marcas ===")
    df_comp = None
    for m in processed_months:
        m_name = m["display_name"]
        summary_df = m["dfs"]["brands_summary"]
        cols_to_use = ["Marca", "Total Retiros", "% Automatizacion"]
        rename_dict = {
            "Total Retiros": f"Total Retiros {m_name}",
            "% Automatizacion": f"% Automatización {m_name}"
        }
        
        if "% Automatizacion Proyectado" in summary_df.columns:
            cols_to_use.append("% Automatizacion Proyectado")
            rename_dict["% Automatizacion Proyectado"] = f"% Automatización Proyectado {m_name}"
            
        df_m_summary = summary_df[cols_to_use].rename(columns=rename_dict)
        if df_comp is None:
            df_comp = df_m_summary
        else:
            df_comp = pd.merge(df_comp, df_m_summary, on="Marca", how="outer")
            
    df_comp = df_comp.fillna(0)
    
    # Calcular evolución entre los últimos dos meses de la lista cronológica
    if len(processed_months) >= 2:
        last_month = processed_months[-1]["display_name"]
        prev_month = processed_months[-2]["display_name"]
        
        # Variación absoluta en puntos porcentuales de la tasa de automatización
        df_comp[f"Variación Tasa Auto ({last_month} vs {prev_month})"] = df_comp[f"% Automatización {last_month}"] - df_comp[f"% Automatización {prev_month}"]
        
        # Crecimiento de volumen de retiros MoM %
        vol_last = df_comp[f"Total Retiros {last_month}"]
        vol_prev = df_comp[f"Total Retiros {prev_month}"]
        df_comp[f"Crecimiento Volumen MoM % ({last_month} vs {prev_month})"] = np.where(
            vol_prev > 0,
            ((vol_last - vol_prev) / vol_prev) * 100,
            0.0
        )
        
    df_comp = df_comp.sort_values("Marca")
    
    print(f"\n=== Escribiendo y formateando '{os.path.basename(excel_out)}' ===")
    with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
        df_comp.to_excel(writer, sheet_name="Resumen Comparativo", index=False)
        workbook = writer.book
        
        # Escribir pestañas de cada mes en orden inverso (más nuevo primero)
        for m in reversed(processed_months):
            ws_name = f"{m['display_name']} 2026"
            ws = workbook.create_sheet(ws_name)
            ws["A1"] = "Resumen de Automatización por Marca"
            m["dfs"]["brands_summary"].to_excel(writer, sheet_name=ws_name, startrow=2, index=False)
            ws["A19"] = "Desglose de Tiempos por Marca y Tipo de Aprobación"
            m["dfs"]["details"].to_excel(writer, sheet_name=ws_name, startrow=20, index=False)
            
        # Pestaña Análisis de Riesgo consolidada
        ws_risk = workbook.create_sheet("Análisis de Riesgo")
        curr_row = 1
        for m in reversed(processed_months):
            ws_risk.cell(row=curr_row, column=1, value=f"Métricas por Nivel de Riesgo - {m['display_name']} 2026")
            m["dfs"]["global_risk"].to_excel(writer, sheet_name="Análisis de Riesgo", startrow=curr_row+1, index=False)
            curr_row += 8
            
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
            
    # Aplicar estilos
    wb = openpyxl.load_workbook(excel_out)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > 30 and cell.column == 1:
                    continue
                max_len = max(max_len, len(val))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        if sheet_name == "Resumen Comparativo":
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = center_align
        elif "2026" in sheet_name:
            for cell in ws[3]:
                cell.font = bold_font
                cell.alignment = center_align
            for cell in ws[21]:
                cell.font = bold_font
                cell.alignment = center_align
        elif sheet_name == "Análisis de Riesgo":
            for r in range(1, ws.max_row + 1):
                val = str(ws.cell(row=r, column=1).value or '')
                if "Métricas por Nivel de Riesgo" in val:
                    ws.cell(row=r, column=1).font = bold_font
                    for cell in ws[r+1]:
                        cell.font = bold_font
                        cell.alignment = center_align
                        
    wb.save(excel_out)
    print("[+] Reporte de Excel formateado y guardado correctamente.")
    
    print("\n=== Actualizando el Dashboard HTML (Inyección JSON) ===")
    if not os.path.exists(html_out):
        print(f"[-] ERROR: Archivo HTML {html_out} no encontrado.")
        return False
        
    with open(html_out, "r", encoding="utf-8") as f:
        html_content = f.read()

    # Encontrar la fecha del último retiro y de última actualización
    import datetime
    last_update_time = datetime.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
    
    last_withdrawal_time = "Desconocido"
    if month_files:
        latest_file = month_files[-1]['file_path']
        latest_csv = latest_file.replace(".xlsx", "_cache.csv")
        try:
            # Intentar cargar desde el archivo CSV cacheado si existe y es más reciente que el archivo de Excel
            use_cache = False
            if os.path.exists(latest_csv):
                excel_mtime = os.path.getmtime(latest_file)
                csv_mtime = os.path.getmtime(latest_csv)
                if csv_mtime > excel_mtime:
                    use_cache = True
            
            if use_cache:
                df_latest = pd.read_csv(latest_csv, usecols=['Fecha Cambio Time'])
            else:
                xl_latest = pd.ExcelFile(latest_file)
                sheet_latest = 'Retiros BD Conexion' if 'Retiros BD Conexion' in xl_latest.sheet_names else xl_latest.sheet_names[0]
                df_latest = xl_latest.parse(sheet_latest, usecols=['Fecha Cambio Time'])
                
            if 'Fecha Cambio Time' in df_latest.columns:
                max_ts = pd.to_datetime(df_latest['Fecha Cambio Time'], errors='coerce').max()
                if pd.notna(max_ts):
                    last_withdrawal_time = max_ts.strftime("%d/%m/%Y %I:%M:%S %p")
        except Exception as e:
            print(f"[-] No se pudo obtener la fecha del último retiro del archivo {latest_file}: {e}")

    # Reemplazar metadatos en el HTML
    metadata_start_key = "const metadata = {"
    m_start_idx = html_content.find(metadata_start_key)
    if m_start_idx != -1:
        m_brace_count = 0
        m_end_idx = -1
        for idx in range(m_start_idx + len("const metadata = ") - 1, len(html_content)):
            char = html_content[idx]
            if char == '{':
                m_brace_count += 1
            elif char == '}':
                m_brace_count -= 1
                if m_brace_count == 0:
                    m_end_idx = idx
                    break
        if m_end_idx != -1:
            metadata_obj = {
                "lastUpdate": last_update_time,
                "lastWithdrawal": last_withdrawal_time
            }
            metadata_injected = json.dumps(metadata_obj, indent=2)
            html_content = html_content[:m_start_idx] + "const metadata = " + metadata_injected + html_content[m_end_idx+1:]
        
    start_key = "const data = {"
    start_idx = html_content.find(start_key)
    if start_idx == -1:
        print("[-] ERROR: No se encontró 'const data = {' en el archivo HTML.")
        return False
        
    brace_count = 0
    end_idx = -1
    for idx in range(start_idx + len("const data = ") - 1, len(html_content)):
        char = html_content[idx]
        if char == '{':
            brace_count += 1
        elif char == '}':
            brace_count -= 1
            if brace_count == 0:
                end_idx = idx
                break
                
    if end_idx == -1:
        print("[-] ERROR: No se cerró correctamente la estructura JSON del HTML.")
        return False
        
    # Calcular MoM Growth para cada marca y para el global en el JSON del Dashboard
    if len(processed_months) > 0:
        first_month = processed_months[0]
        first_month["json_data"]["global"]["mom_vol_growth"] = 0.0
        first_month["json_data"]["global"]["mom_pct_diff"] = 0.0
        for b in first_month["json_data"]["brands_summary"]:
            b["mom_vol_growth"] = 0.0
            b["mom_pct_diff"] = 0.0
            b["mom_proj_diff"] = 0.0
            
    for i in range(1, len(processed_months)):
        curr_month = processed_months[i]
        prev_month = processed_months[i-1]
        
        # 1. Global MoM
        curr_glob = curr_month["json_data"]["global"]
        prev_glob = prev_month["json_data"]["global"]
        
        v_curr_g = curr_glob["total_retiros"]
        v_prev_g = prev_glob["total_retiros"]
        curr_glob["mom_vol_growth"] = float(((v_curr_g - v_prev_g) / v_prev_g * 100) if v_prev_g > 0 else 0.0)
        curr_glob["mom_pct_diff"] = float(curr_glob["pct_automatizacion"] - prev_glob["pct_automatizacion"])
        
        # 2. Brand MoM
        prev_brands = {b["Marca"]: b for b in prev_month["json_data"]["brands_summary"]}
        for curr_brand_data in curr_month["json_data"]["brands_summary"]:
            brand_name = curr_brand_data["Marca"]
            prev_brand_data = prev_brands.get(brand_name)
            if prev_brand_data:
                v_curr = curr_brand_data["Total Retiros"]
                v_prev = prev_brand_data["Total Retiros"]
                curr_brand_data["mom_vol_growth"] = float(((v_curr - v_prev) / v_prev * 100) if v_prev > 0 else 0.0)
                curr_brand_data["mom_pct_diff"] = float(curr_brand_data["% Automatizacion"] - prev_brand_data["% Automatizacion"])
                curr_brand_data["mom_proj_diff"] = float(curr_brand_data["% Automatizacion Proyectado"] - prev_brand_data["% Automatizacion Proyectado"])
            else:
                curr_brand_data["mom_vol_growth"] = 0.0
                curr_brand_data["mom_pct_diff"] = 0.0
                curr_brand_data["mom_proj_diff"] = 0.0

    # Construir diccionario consolidado
    js_data = {}
    for m in processed_months:
        js_data[m["month_name"]] = clean_nan(m["json_data"])
        
    json_injected = json.dumps(js_data, indent=2)
    new_html = html_content[:start_idx] + "const data = " + json_injected + ";" + html_content[end_idx+1:]
    
    # Pre-render static HTML table rows for metrics-table to guarantee instant display
    try:
        def _prerender_duration(hrs):
            if hrs is None or hrs != hrs:
                return 'N/D'
            if hrs < 1.0:
                mins = round(hrs * 60)
                return f"{mins}m"
            hours = int(hrs)
            mins = round((hrs - hours) * 60)
            return f"{hours}h {mins}m" if mins > 0 else f"{hours}h"

        tableAgg = {}
        for m_name, m_val in js_data.items():
            for r in m_val.get("daily_data", []):
                count = r.get("count") or r.get("total") or 0
                isAuto = (r.get("Tipo Aprobacion_Norm") == "Automatico" or r.get("aprobacion") == "Automatico")
                count_auto = count if isAuto else 0
                count_manual = count - count_auto
                marca = r.get("Marca")
                if count_auto > 0:
                    keyAuto = f"{marca}__Automatico"
                    if keyAuto not in tableAgg:
                        tableAgg[keyAuto] = {"Marca": marca, "Tipo": "Automatico", "count": 0, "sum_ca": 0, "count_ca": 0, "sum_cp": 0, "count_cp": 0, "sum_ap": 0, "count_ap": 0}
                    tableAgg[keyAuto]["count"] += count_auto
                    tableAgg[keyAuto]["sum_ca"] += (r.get("sum_ca_min", 0) * (count_auto / count)) if count > 0 else 0
                    tableAgg[keyAuto]["count_ca"] += (r.get("count_ca", 0) * (count_auto / count)) if count > 0 else 0
                    tableAgg[keyAuto]["sum_cp"] += (r.get("sum_cp_min", 0) * (count_auto / count)) if count > 0 else 0
                    tableAgg[keyAuto]["count_cp"] += (r.get("count_cp", 0) * (count_auto / count)) if count > 0 else 0
                    tableAgg[keyAuto]["sum_ap"] += (r.get("sum_ap_min", 0) * (count_auto / count)) if count > 0 else 0
                    tableAgg[keyAuto]["count_ap"] += (r.get("count_ap", 0) * (count_auto / count)) if count > 0 else 0
                    
                if count_manual > 0:
                    keyManual = f"{marca}__Manual"
                    if keyManual not in tableAgg:
                        tableAgg[keyManual] = {"Marca": marca, "Tipo": "Manual", "count": 0, "sum_ca": 0, "count_ca": 0, "sum_cp": 0, "count_cp": 0, "sum_ap": 0, "count_ap": 0}
                    tableAgg[keyManual]["count"] += count_manual
                    tableAgg[keyManual]["sum_ca"] += (r.get("sum_ca_min", 0) * (count_manual / count)) if count > 0 else 0
                    tableAgg[keyManual]["count_ca"] += (r.get("count_ca", 0) * (count_manual / count)) if count > 0 else 0
                    tableAgg[keyManual]["sum_cp"] += (r.get("sum_cp_min", 0) * (count_manual / count)) if count > 0 else 0
                    tableAgg[keyManual]["count_cp"] += (r.get("count_cp", 0) * (count_manual / count)) if count > 0 else 0
                    tableAgg[keyManual]["sum_ap"] += (r.get("sum_ap_min", 0) * (count_manual / count)) if count > 0 else 0
                    tableAgg[keyManual]["count_ap"] += (r.get("count_ap", 0) * (count_manual / count)) if count > 0 else 0

        sortedRows = sorted(tableAgg.values(), key=lambda x: (x["Marca"], x["Tipo"]))
        rows_html = []
        for r in sortedRows:
            badgeClass = "badge-auto" if r["Tipo"] == "Automatico" else "badge-manual"
            caHrs = (r["sum_ca"] / r["count_ca"] / 60.0) if r["count_ca"] > 0 else None
            cpHrs = (r["sum_cp"] / r["count_cp"] / 60.0) if r["count_cp"] > 0 else None
            apHrs = (r["sum_ap"] / r["count_ap"] / 60.0) if r["count_ap"] > 0 else None
            caStyle = 'style="color: #f87171; font-weight: 500;"' if (caHrs is not None and caHrs > 1.0) else ''
            row_str = f'''                <tr>
                    <td><strong>{r['Marca']}</strong></td>
                    <td><span class="badge {badgeClass}">{r['Tipo']}</span></td>
                    <td>{r['count']:,}</td>
                    <td {caStyle}><div>Prom: {_prerender_duration(caHrs)}</div></td>
                    <td><div>Prom: {_prerender_duration(cpHrs)}</div></td>
                    <td><div>Prom: {_prerender_duration(apHrs)}</div></td>
                </tr>'''
            rows_html.append(row_str)

        table_body_content = "\n" + "\n".join(rows_html) + "\n            "
        table_pattern = r'(<table id="metrics-table"[\s\S]*?<tbody>)([\s\S]*?)(<\/tbody>)'
        new_html = re.sub(table_pattern, r"\1" + table_body_content + r"\3", new_html)
    except Exception as e_prerender:
        print(f"[-] Warning: Pre-rendering table failed: {e_prerender}")

    with open(html_out, "w", encoding="utf-8") as f:
        f.write(new_html)
        
    index_out = os.path.join(dir_path, "index.html")
    with open(index_out, "w", encoding="utf-8") as f:
        f.write(new_html)
        
    print(f"[+] Dashboard_Efectividad.html actualizado exitosamente con metadatos (Actualización: {last_update_time}, Último Retiro: {last_withdrawal_time}) y todos los meses.")
    return True

def run_pipeline():
    if USERNAME == "tu_usuario" or PASSWORD == "tu_contraseña":
        print("[-] Por favor, edita este script e introduce tu usuario y contraseña reales de MicroStrategy.")
        return

    session = requests.Session()
    session.verify = False

    print("[*] 1. Intentando autenticación en MicroStrategy REST API...")
    login_url = f"{API_URL}/auth/login"
    payload = {
        "username": USERNAME,
        "password": PASSWORD,
        "loginMode": 1
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    try:
        response = session.post(login_url, json=payload, headers=headers)
        if response.status_code not in [200, 204]:
            print(f"[-] Error de autenticación. Código: {response.status_code}")
            print(response.text)
            return

        auth_token = response.headers.get("X-MSTR-AuthToken")
        if not auth_token:
            print("[-] No se recibió el token de autenticación (X-MSTR-AuthToken).")
            return

        print("[+] Autenticación exitosa! Token obtenido.")
        session.headers.update({
            "X-MSTR-AuthToken": auth_token,
            "Accept": "application/json"
        })

        print("\n[*] 2. Obteniendo lista de Proyectos disponibles...")
        projects_url = f"{API_URL}/projects"
        proj_resp = session.get(projects_url)
        
        if proj_resp.status_code != 200:
            print(f"[-] No se pudo obtener la lista de proyectos. Código: {proj_resp.status_code}")
            return

        projects = proj_resp.json()
        project_id = None
        for p in projects:
            p_name = p.get("name", "")
            p_id = p.get("id", "")
            if "virtualsoft" in p_name.lower():
                project_id = p_id
                print(f"    [!] Proyecto seleccionado: '{p_name}' | ID: '{p_id}'")
                break

        if not project_id and projects:
            project_id = projects[0].get("id")
            print(f"    [!] Usando primer proyecto: '{projects[0].get('name')}'")

        if not project_id:
            print("[-] No se pudo determinar ningún ID de proyecto.")
            return

        session.headers.update({"X-MSTR-ProjectID": project_id})

        print(f"\n[*] 3. Creando instancia del Reporte de MicroStrategy para ID {DOSSIER_ID} con filtro de fecha...")
        # Filtro incremental: descargar solo los últimos 7 días para rapidez extrema
        import datetime
        now = datetime.datetime.now()
        start_date = now - datetime.timedelta(days=7)
        filter_date_str = start_date.strftime("%Y-%m-%d 00:00:00")
        print(f"    [+] [Modo Incremental] Aplicando filtro de fecha (últimos 7 días): Fecha Cambio Time >= {filter_date_str}")
        
        payload = {
            "viewFilter": {
                "operator": "Between",
                "operands": [
                    {
                        "type": "form",
                        "attribute": {
                            "id": "A632D4914460225A909A94A4A411506C" # Fecha Cambio Time
                        },
                        "form": {
                            "id": "CCFBE2A5EADB4F50941FB879CCF1721C" # DESC Form
                        }
                    },
                    {
                        "type": "constant",
                        "dataType": "TimeStamp",
                        "value": filter_date_str
                    },
                    {
                        "type": "constant",
                        "dataType": "TimeStamp",
                        "value": "2030-12-31 23:59:59" # Fecha límite lejana
                    }
                ]
            }
        }

        report_url = f"{API_URL}/reports/{DOSSIER_ID}/instances"
        resp = session.post(report_url, json=payload)
        if resp.status_code not in [200, 201]:
            print(f"[-] Error al crear la instancia del reporte: {resp.status_code}")
            print(resp.text[:500])
            return

        instance_data = resp.json()
        instance_id = instance_data.get("instanceId")
        result_obj = instance_data.get("result", {})
        definition = result_obj.get("definition", {})
        paging = result_obj.get("data", {}).get("paging", {})
        total_rows = paging.get("total", 0)

        print(f"[+] Instancia creada con éxito. ID: {instance_id}")
        print(f"[+] Total de filas reportadas por el servidor: {total_rows}")

        # Extraer nombres de atributos
        attributes_def = definition.get("grid", {}).get("rows", [])
        if not attributes_def:
            attributes_def = definition.get("attributes", [])
        attribute_names = [attr.get("name") for attr in attributes_def]
        print(f"[+] Atributos del Reporte: {attribute_names}")

        # Descarga paginada
        all_parsed_rows = []
        limit = 50000
        offset = 0
        start_time = time.time()

        print("\n[*] 4. Iniciando descarga paginada en lotes de 50k...")
        while offset < total_rows:
            page_start = time.time()
            page_url = f"{API_URL}/reports/{DOSSIER_ID}/instances/{instance_id}?offset={offset}&limit={limit}"
            
            # Intentar la descarga con reintentos
            max_retries = 5
            retry_delay = 5
            page_resp = None
            
            for retry in range(max_retries):
                try:
                    page_resp = session.get(page_url)
                    if page_resp.status_code == 200:
                        break
                    else:
                        print(f"\n[!] Advertencia: Intento {retry+1} falló para offset {offset} con código {page_resp.status_code}. Reintentando en {retry_delay}s...")
                except Exception as e:
                    print(f"\n[!] Excepción en intento {retry+1} para offset {offset}: {e}. Reintentando en {retry_delay}s...")
                
                time.sleep(retry_delay)
                retry_delay *= 2  # Backoff exponencial
                
            if page_resp is None or page_resp.status_code != 200:
                print(f"\n[-] ERROR CRÍTICO: No se pudo descargar la página con offset {offset} tras {max_retries} intentos.")
                return
                
            page_data = page_resp.json()
            page_result = page_data.get("result", {})
            page_root = page_result.get("data", {}).get("root", {})
            
            page_rows = parse_mstr_node(page_root, attribute_names)
            all_parsed_rows.extend(page_rows)
            
            page_elapsed = time.time() - page_start
            print(f"    - Lote offset {offset:6d} | Descargadas {len(page_rows):5d} filas en {page_elapsed:.2f}s (Acumulado: {len(all_parsed_rows):6d})")
            
            offset += limit

        total_elapsed = time.time() - start_time
        print(f"\n[+] ¡Descarga completada con éxito!")
        print(f"[+] Total de filas en bruto descargadas: {len(all_parsed_rows)}")
        print(f"[+] Tiempo de descarga y parseo: {total_elapsed:.2f}s")

        if not all_parsed_rows:
            print("[-] No se descargaron registros del reporte.")
            return

        # Convertir a DataFrame y agrupar por meses detectados para guardarlos por separado
        df_raw = pd.DataFrame(all_parsed_rows)
        temp_dates = pd.to_datetime(df_raw['Fecha Cambio Time'], errors='coerce')
        df_raw['temp_month_num'] = temp_dates.dt.month
        df_raw['temp_year'] = temp_dates.dt.year
        
        # Encontrar los meses únicos en la descarga y ordenarlos cronológicamente
        unique_months_df = df_raw[['temp_year', 'temp_month_num']].dropna().drop_duplicates()
        unique_months_df = unique_months_df.sort_values(by=['temp_year', 'temp_month_num'])
        unique_months = unique_months_df.values
        
        month_num_to_name = {
            1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
            7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
        }
        
        print("\n[*] 5. Procesando data descargada por meses...")
        for yr, m_num in unique_months:
            m_name = month_num_to_name.get(int(m_num), "desconocido")
            display_name = m_name.capitalize()
            
            target_excel_name = f"Retiros {display_name}.xlsx"
            target_excel_path = os.path.join(dir_path, target_excel_name)
            
            # EVITAR RE-ESCRITURA DE HISTÓRICOS CERRADOS:
            # Los meses anteriores al mes previo (ej: Enero, Febrero, Marzo, Abril si estamos en Junio)
            # no se vuelven a guardar en Excel si ya existen en disco, lo que ahorra más del 90% del tiempo de ejecución.
            import datetime
            current_month = datetime.datetime.now().month
            is_closed_month = int(m_num) < (current_month - 1)
            
            if is_closed_month and os.path.exists(target_excel_path) and os.path.getsize(target_excel_path) > 0:
                print(f"\n    - [Omisión] {display_name} {int(yr)} es un mes histórico cerrado y ya existe. Se omite escritura.")
                continue
                
            # Filtrar registros correspondientes a este mes específico
            df_m_raw = df_raw[(df_raw['temp_year'] == yr) & (df_raw['temp_month_num'] == m_num)].copy()
            df_m_raw = df_m_raw.drop(columns=['temp_month_num', 'temp_year'])
            
            print(f"\n    - Lote detectado para: {display_name} {int(yr)} ({len(df_m_raw)} filas nuevas descargadas)")
            
            print(f"    [*] Limpiando lote incremental...")
            df_new_clean = clean_mstr_dataframe(df_m_raw)
            
            # Cargar data histórica del mes actual si existe para fusionar localmente
            if os.path.exists(target_excel_path):
                print(f"        [+] Archivo existente detectado. Fusionando incrementalmente...")
                try:
                    target_csv_path = target_excel_path.replace(".xlsx", "_cache.csv")
                    if os.path.exists(target_csv_path):
                        df_old = pd.read_csv(target_csv_path)
                    else:
                        df_old = pd.read_excel(target_excel_path)
                    
                    # Asegurar Id Retiro como tipo uniforme para la comparación
                    df_old['Id Retiro_str'] = df_old['Id Retiro'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                    df_new_clean['Id Retiro_str'] = df_new_clean['Id Retiro'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                    
                    # Remover del conjunto antiguo cualquier ID que esté en el conjunto nuevo
                    df_old_filtered = df_old[~df_old['Id Retiro_str'].isin(df_new_clean['Id Retiro_str'])].copy()
                    
                    # Limpiar columnas temporales
                    df_old_filtered = df_old_filtered.drop(columns=['Id Retiro_str'], errors='ignore')
                    df_new_clean = df_new_clean.drop(columns=['Id Retiro_str'], errors='ignore')
                    
                    # Concatenar para unir
                    df_clean = pd.concat([df_old_filtered, df_new_clean], ignore_index=True)
                    print(f"        [+] Mezcla completada: {len(df_old)} existentes + {len(df_new_clean)} nuevos -> {len(df_clean)} filas finales.")
                except Exception as merge_err:
                    print(f"        [-] Error al fusionar, se usará solo el lote nuevo: {merge_err}")
                    df_clean = df_new_clean
            else:
                print(f"        [!] No existe archivo local previo. Guardando lote de descarga como base.")
                df_clean = df_new_clean
            
            # Encontrar y mover actualizaciones de meses anteriores
            df_clean['Id Retiro_str'] = df_clean['Id Retiro'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
            
            for prev_m_num in range(1, int(m_num)):
                prev_m_name = month_num_to_name.get(prev_m_num)
                if prev_m_name:
                    hist_path = os.path.join(dir_path, f"Retiros {prev_m_name.capitalize()}.xlsx")
                    hist_csv_path = hist_path.replace(".xlsx", "_cache.csv")
                    if os.path.exists(hist_path):
                        print(f"        - Cargando IDs históricos de '{os.path.basename(hist_path)}' (usando caché)...")
                        try:
                            # Intentar cargar desde caché si existe y es más reciente
                            use_hist_cache = False
                            if os.path.exists(hist_csv_path):
                                excel_mtime = os.path.getmtime(hist_path)
                                csv_mtime = os.path.getmtime(hist_csv_path)
                                if csv_mtime > excel_mtime:
                                    use_hist_cache = True
                                    
                            if use_hist_cache:
                                df_hist = pd.read_csv(hist_csv_path)
                            else:
                                df_hist = pd.read_excel(hist_path)
                                df_hist.to_csv(hist_csv_path, index=False)
                                
                            if 'Id Retiro' in df_hist.columns:
                                df_hist['Id Retiro_str'] = df_hist['Id Retiro'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                                
                                # Encontrar la intersección de IDs
                                overlapping_ids = set(df_clean['Id Retiro_str']).intersection(set(df_hist['Id Retiro_str'].dropna()))
                                if overlapping_ids:
                                    print(f"        [!] Se detectaron {len(overlapping_ids)} retiros que estaban en '{os.path.basename(hist_path)}' pero tienen fecha de cambio en {display_name}.")
                                    print(f"            - Moviendo estos registros a '{target_excel_name}'...")
                                    
                                    # Remover del archivo histórico
                                    df_hist_new = df_hist[~df_hist['Id Retiro_str'].isin(overlapping_ids)].copy()
                                    df_hist_new = df_hist_new.drop(columns=['Id Retiro_str'], errors='ignore')
                                    
                                    # Guardar archivo histórico actualizado
                                    df_hist_new.to_excel(hist_path, sheet_name="Retiros BD Conexion", index=False)
                                    df_hist_new.to_csv(hist_csv_path, index=False)
                                    print(f"            [+] Archivo histórico '{os.path.basename(hist_path)}' y su caché actualizados con éxito.")
                                
                        except Exception as ex_hist:
                            print(f"        - Error al actualizar histórico: {ex_hist}")
            
            df_clean = df_clean.drop(columns=['Id Retiro_str'], errors='ignore')
            if os.path.exists(target_excel_path):
                backup_path = target_excel_path.replace(".xlsx", "_Prev.xlsx")
                try:
                    if os.path.exists(backup_path):
                        os.remove(backup_path)
                    os.rename(target_excel_path, backup_path)
                    print(f"        - Copia de respaldo temporal creada en '{backup_path}'")
                except Exception as e:
                    print(f"        - Advertencia al crear respaldo: {e}")
                    
            df_clean.to_excel(target_excel_path, sheet_name="Retiros BD Conexion", index=False)
            try:
                df_clean.to_csv(target_excel_path.replace(".xlsx", "_cache.csv"), index=False)
            except Exception as e:
                print(f"        [-] Advertencia al crear caché CSV para {target_excel_name}: {e}")
            print(f"    [+] Archivo '{target_excel_name}' y su caché guardados correctamente.")

        # Ejecutar la tubería de análisis y reportes para todos los archivos mensuales
        success = run_reporting_pipeline()
        if success:
            print("\n======================================================================")
            print("¡ACTUALIZACIÓN COMPLETADA EXITOSAMENTE EXTREMO A EXTREMO!")
            print("- Todos los meses se actualizaron correctamente con la descarga completa.")
            print(f"- Se procesaron todos los archivos de meses en la carpeta.")
            print("- El consolidado 'Reporte_Efectividad_Automatizacion.xlsx' fue actualizado.")
            print("- El tablero 'Dashboard_Efectividad.html' fue inyectado y dinamizado.")
            print("======================================================================")
        else:
            print("\n[-] Ocurrió un error en la tubería de reportes.")

    except Exception as e:
        print(f"\n[-] Error durante la conexión y procesamiento: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    run_pipeline()
